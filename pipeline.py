# -*- coding: utf-8 -*-
"""
PIPELINE TUYỂN DỤNG GIẢNG VIÊN UEH — chạy 1 lệnh
=================================================
Quy trình ETL: gộp 8 file Excel thô trong ``data/`` thành 1 file đúng định dạng
``Duthaokehoachtuyendung2026-2030.xlsx`` (các sheet "có sử dụng"), rồi gọi
``build_data`` để tính định biên và xuất ``data.js`` / ``data.json`` cho dashboard.

CÁCH DÙNG
    1. Bỏ 8 file Excel nguồn vào thư mục  data/
    2. Chạy:  python3 pipeline.py
    3. Kết quả:
         - data/Duthaokehoach_merged.xlsx   (file gộp)
         - data.js / data.json              (dữ liệu cho index.html)

Script TỰ ĐỘNG NHẬN DIỆN từng file nguồn theo cấu trúc cột (không phụ thuộc tên
file hay thứ tự), nên chỉ cần thả 8 file vào là chạy được.

Các loại dữ liệu nguồn cần có (file thừa sẽ bị bỏ qua):
    hr           Danh sách giảng viên   -> sheet 'Quy mô giảng viên'
    demand       Danh sách môn học      -> sheet 'Phụ lục 2. DS môn học 2025'
    course_unit  Môn học -> Khoa phụ trách   (ghép vào Phụ lục 2)
    ug           Quy mô sinh viên ĐH    -> sheet 'Quy mô sinh viên'
    ctdt         Chương trình đào tạo   -> sheet 'Chuongtrinhdaotao'
    sdh          Học viên/NCS SĐH       -> sheet 'Quy mô SĐH'
    map          Đơn vị giảng dạy       -> sheet 'Đơn vị giảng dạy'
"""
from __future__ import annotations

import os
import re
import sys
import glob
import datetime
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from typing import Optional


def _ensure_utf8_stdout() -> None:
    """Đảm bảo stdout ghi UTF-8 (terminal Windows hay mặc định cp1252)."""
    try:
        enc = (sys.stdout.encoding or "").lower()
        if not enc.startswith("utf"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


_ensure_utf8_stdout()

import warnings

import openpyxl
from openpyxl import Workbook

# openpyxl cảnh báo "Data Validation extension" với 1 số file UEH — vô hại
warnings.filterwarnings("ignore", message="Data Validation extension is not supported")

import build_data

# =========================================================
# CẤU HÌNH
# =========================================================
BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "data")
MERGED_PATH = os.path.join(DATA_DIR, "Duthaokehoach_merged.xlsx")

# Tên sheet đích — phải khớp build_data.SHEETS
SHEET_HR = "Quy mô giảng viên"
SHEET_DEMAND = "Phụ lục 2. DS môn học 2025"
SHEET_UG = "Quy mô sinh viên"
SHEET_PG = "Quy mô SĐH"
SHEET_MAP = "Đơn vị giảng dạy"
SHEET_CTDT = "Chuongtrinhdaotao"

# Đơn vị hành chính thuần — KHÔNG tính định biên (dù có dạy vài học phần)
ADMIN_UNITS = ("ban giám hiệu", "văn phòng", "phòng chăm sóc")
# Trạng thái học viên SĐH được tính vào "quy mô đang đào tạo"
SDH_ACTIVE_STATUS = "đang học"

PREFIX_RE = re.compile(r"^[A-ZĐ]{2,8}\s*-\s*")
DIGITS_RE = re.compile(r"\D")
TOTAL_KEYWORDS = {"CỘNG", "TỔNG", "TỔNG CỘNG", "TONG", "CONG"}


# =========================================================
# HELPER: chuẩn hóa & chuyển kiểu
# =========================================================
def norm(v) -> str:
    """Strip + Unicode NFC (đồng nhất dấu tổ hợp/precomposed)."""
    if v is None:
        return ""
    return unicodedata.normalize("NFC", str(v).strip())


def to_int(v) -> Optional[int]:
    s = norm(v).replace(",", "")
    if not s or s.lower() in ("none", "nan"):
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def to_num(v):
    """Trả về int/float; rỗng -> 0 (dùng cho cột số liệu cộng dồn)."""
    s = norm(v).replace(",", "")
    if not s or s.lower() in ("none", "nan"):
        return 0
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return 0


def to_date(v) -> Optional[datetime.datetime]:
    if isinstance(v, datetime.datetime):
        return v
    s = norm(v)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except (ValueError, TypeError):
            continue
    return None


def to_bool(v) -> Optional[bool]:
    s = norm(v).lower()
    if s in ("true", "1", "nam", "male", "x"):
        return True
    if s in ("false", "0", "nu", "nữ", "female"):
        return False
    return None


def col_index(header, *names) -> Optional[int]:
    """Chỉ số cột khớp đầu tiên (ưu tiên khớp đúng, sau đó khớp chứa)."""
    low = [h.lower() for h in header]
    for n in names:
        n = n.lower()
        for i, h in enumerate(low):
            if h == n:
                return i
    for n in names:
        n = n.lower()
        for i, h in enumerate(low):
            if n in h:
                return i
    return None


# =========================================================
# ĐỌC EXCEL (mở mỗi workbook 1 lần, có cache)
# =========================================================
@lru_cache(maxsize=64)
def _load_sheet(path: str, sheet: str):
    """Đọc toàn bộ 1 sheet -> tuple[tuple] (cache theo path+sheet)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows = tuple(wb[sheet].iter_rows(values_only=True))
    finally:
        wb.close()
    return rows


def read_sheet(ref: "SourceRef"):
    """Trả về (header:list[str], data_rows:list[tuple]) theo header_row của ref."""
    rows = _load_sheet(ref.path, ref.sheet)
    if ref.header_row > len(rows):
        return [], []
    header = [norm(c) for c in rows[ref.header_row - 1]]
    return header, list(rows[ref.header_row:])


def preview_workbook(path: str, n: int = 8):
    """Mở workbook 1 lần, trả về {sheet: [n dòng đầu]} để nhận diện (rẻ)."""
    out = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = []
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i >= n:
                    break
                rows.append(r)
            out[ws.title] = rows
    finally:
        wb.close()
    return out


# =========================================================
# NHẬN DIỆN FILE NGUỒN
# =========================================================
@dataclass(frozen=True)
class SourceRef:
    path: str
    sheet: str
    header_row: int   # 1-based

    @property
    def filename(self) -> str:
        return os.path.basename(self.path)


def _classify_sheet(cells_by_row):
    """Nhận diện vai trò của 1 sheet từ các dòng đầu -> (role, header_row) hoặc None."""
    for ri, row in enumerate(cells_by_row):
        low = [norm(c).lower() for c in row]
        hrow = ri + 1
        s = set(low)

        if {"mã quản lý", "học vị", "tình trạng làm việc", "chức danh", "họ tên"} <= s:
            return "hr", hrow
        if "mã hp" in s and ("mã người phụ trách" in s or "tên người phụ trách" in s) and "khoa" in s:
            return "course_unit", hrow
        if "mã hp" in s and "tổng số tiết" in s and "số tín chỉ" in s and "khoa" not in s:
            return "demand", hrow
        if "isb - chính quy" in s and "cộng" in s:
            return "ug", hrow
        if "mã quản lý" in s and "đơn vị giảng dạy" in s and len(row) <= 4:
            return "map", hrow
        if "mã chương trình đào tạo" in s and "trạng thái học" in s:
            return "sdh", hrow
        if "mã chương trình" in s and "khoa/viện" in s and "trình độ" in s:
            return "ctdt", hrow
    return None


def discover_sources(data_dir: str):
    """Quét data/*.xlsx, nhận diện vai trò từng file. -> (sources, unused, errors)."""
    if not os.path.isdir(data_dir):
        raise FileNotFoundError(f"Khong tim thay thu muc du lieu: {data_dir}")

    files = sorted(
        f for f in glob.glob(os.path.join(data_dir, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
        and os.path.abspath(f) != os.path.abspath(MERGED_PATH)
    )
    if not files:
        raise FileNotFoundError(f"Khong co file .xlsx nao trong {data_dir}")

    sources: dict[str, SourceRef] = {}
    unused, errors = [], []

    for path in files:
        try:
            preview = preview_workbook(path)
        except Exception as e:                       # file hỏng/đang mở/khóa
            errors.append((os.path.basename(path), str(e)))
            continue

        matched_any = False
        for sheet, rows in preview.items():
            hit = _classify_sheet(rows)
            if hit:
                role, hrow = hit
                sources.setdefault(role, SourceRef(path, sheet, hrow))
                matched_any = True
        if not matched_any:
            unused.append(path)

    return sources, unused, errors


# =========================================================
# CHUẨN HÓA TÊN ĐƠN VỊ:  "Khoa X" -> "PREFIX - Khoa X"
# =========================================================
def build_canonical_map(course_unit_ref: SourceRef, map_ref: SourceRef) -> dict:
    """Bảng tra: tên đơn vị (bỏ prefix, lowercase) -> tên canonical có prefix."""
    canon: dict[str, str] = {}
    pools = []

    ch, crows = read_sheet(course_unit_ref)
    ci = col_index(ch, "Khoa")
    if ci is not None:
        pools += [norm(r[ci]) for r in crows if ci < len(r)]

    mh, mrows = read_sheet(map_ref)
    mi = col_index(mh, "Đơn vị giảng dạy")
    if mi is not None:
        pools += [norm(r[mi]) for r in mrows if mi < len(r)]

    for name in pools:
        if name and PREFIX_RE.match(name):
            plain = PREFIX_RE.sub("", name).strip().lower()
            canon.setdefault(plain, name)
    return canon


def normalize_unit(name, canon: dict) -> str:
    name = norm(name)
    if not name or PREFIX_RE.match(name):     # rỗng hoặc đã có prefix
        return name
    return canon.get(name.lower(), name)


def is_teaching_unit(name: str) -> bool:
    """Đơn vị có tính định biên (Khoa/Viện/Trung tâm/Phòng dạy học), trừ hành chính."""
    plain = PREFIX_RE.sub("", name).strip().lower()
    if any(a in plain for a in ADMIN_UNITS):
        return False
    return bool(PREFIX_RE.match(name)) or plain.startswith(
        ("khoa", "viện", "trung tâm", "phòng"))


# =========================================================
# DỰNG TỪNG SHEET "CÓ SỬ DỤNG"
# =========================================================
def build_hr(ref: SourceRef):
    header, rows = read_sheet(ref)
    out_cols = ['STT', 'Mã quản lý', 'Họ tên', 'CMND', 'Ngày sinh', 'Giới tính', 'Quốc tịch',
                'Email', 'Điện thoại', 'Chức vụ', 'Đơn vị', 'Tổ bộ môn', 'Học hàm', 'Học vị',
                'Chuyên ngành', 'Nơi đào tạo', 'Quốc gia đào tạo', 'Năm tốt nghiệp', 'Mã ngạch',
                'Chức danh', 'Tình trạng làm việc', 'Tên phân hiệu']
    cmap = {h.lower(): i for i, h in enumerate(header)}

    def g(row, name):
        i = cmap.get(name.lower())
        return row[i] if (i is not None and i < len(row)) else None

    # (tên cột nguồn, hàm chuyển kiểu) theo đúng thứ tự out_cols (trừ STT tự đánh)
    spec = [
        ('Mã quản lý', to_int), ('Họ tên', norm), ('CMND', to_int), ('Ngày sinh', to_date),
        ('Giới tính', to_bool), ('Quốc tịch', norm), ('Email', norm), ('Điện thoại', to_int),
        ('Chức vụ', norm), ('Đơn vị', norm), ('Tổ bộ môn', norm), ('Học hàm', norm),
        ('Học vị', norm), ('Chuyên ngành', norm), ('Nơi đào tạo', norm),
        ('Quốc gia đào tạo', norm), ('Năm tốt nghiệp', to_int), ('Mã ngạch', norm),
        ('Chức danh', norm), ('Tình trạng làm việc', norm), ('Tên phân hiệu', norm),
    ]
    out = [out_cols]
    n = 0
    for r in rows:
        if not norm(g(r, "Họ tên")) and not norm(g(r, "Mã quản lý")):
            continue
        n += 1
        vals = [n]
        for name, conv in spec:
            v = conv(g(r, name))
            vals.append(v if v != "" else None)
        out.append(vals)
    return out, n


def build_demand(demand_ref: SourceRef, course_unit_ref: SourceRef, canon: dict):
    dh, drows = read_sheet(demand_ref)
    di = {k: col_index(dh, label) for k, label in [
        ("ma", "Mã HP"), ("ten", "Tên HP"), ("bac", "Bậc đào tạo"),
        ("lhp", "Số LHP"), ("tc", "Số tín chỉ"), ("tiet", "Tổng số tiết")]}

    ch, crows = read_sheet(course_unit_ref)
    ci_ma, ci_khoa = col_index(ch, "Mã HP"), col_index(ch, "Khoa")
    ma2khoa: dict[str, str] = {}
    for r in crows:
        ma = norm(r[ci_ma]) if ci_ma is not None and ci_ma < len(r) else ""
        khoa = normalize_unit(r[ci_khoa] if ci_khoa is not None and ci_khoa < len(r) else "", canon)
        if ma and khoa:
            ma2khoa.setdefault(ma, khoa)

    teaching = {v.lower() for v in ma2khoa.values() if is_teaching_unit(v)}

    out = [['Mã HP', 'Tên HP', 'Bậc đào tạo', 'Số LHP', 'Số tín chỉ', 'Tổng số tiết',
            'Ghi chú', 'Đơn vị phụ trách']]
    n = 0
    for r in drows:
        ma = norm(r[di["ma"]]) if di["ma"] is not None else ""
        if not ma:
            continue
        khoa = ma2khoa.get(ma, "#N/A")
        n += 1
        out.append([
            ma,
            norm(r[di["ten"]]) or None if di["ten"] is not None else None,
            norm(r[di["bac"]]) or None if di["bac"] is not None else None,
            to_num(r[di["lhp"]]) if di["lhp"] is not None else None,
            to_num(r[di["tc"]]) if di["tc"] is not None else None,
            to_num(r[di["tiet"]]) if di["tiet"] is not None else None,
            "Tính định biên" if khoa.lower() in teaching else "",
            khoa,
        ])
    return out, n


def build_ug(ref: SourceRef, canon: dict):
    header, rows = read_sheet(ref)
    base_cols = ['Tên chương trình đào tạo', 'Khoa/Viện quản lý', 'ISB - Chính quy',
                 'Chính quy', 'Văn bằng 2 chính quy', 'Liên thông chính quy',
                 'Văn bằng 1 - Vừa làm vừa học', 'Văn bằng 2 - Vừa làm vừa học',
                 'Liên thông cao đẳng - Vừa làm vừa học',
                 'Liên thông trung cấp - Vừa làm vừa học', 'CỘNG']
    cidx = [col_index(header, c) for c in base_cols]
    cq_cols = base_cols[2:6]                 # ISB-CQ ... Liên thông chính quy
    vlvh_cols = base_cols[6:10]              # 4 cột VLVH
    out = [base_cols + ['tong ĐHCQ', 'VLVH']]
    n = 0
    for r in rows:
        ten = norm(r[cidx[0]]) if cidx[0] is not None else ""
        khoa = norm(r[cidx[1]]) if (cidx[1] is not None and cidx[1] < len(r)) else ""
        # bỏ dòng trống và dòng tổng cuối bảng (vd 'CỘNG' không có Khoa/Viện)
        if not ten or ten.upper() in TOTAL_KEYWORDS or not khoa:
            continue
        n += 1
        vals = []
        for j, ci in enumerate(cidx):
            v = r[ci] if (ci is not None and ci < len(r)) else None
            if j >= 2:
                vals.append(to_num(v))
            elif j == 1:
                vals.append(normalize_unit(v, canon))
            else:
                vals.append(norm(v) or None)
        rowmap = dict(zip(base_cols, vals))
        out.append(vals + [sum(rowmap[c] for c in cq_cols),
                           sum(rowmap[c] for c in vlvh_cols)])
    return out, n


def build_pg(sdh_ref: SourceRef, ctdt_ref: SourceRef, canon: dict):
    """Quy mô SĐH = đếm học viên/NCS 'đang học' theo mã ngành (7 số đầu mã CTĐT)."""
    # CTDT: mã ngành 7 ký tự -> (tên ngành, đơn vị canonical) + số chương trình
    ch, crows = read_sheet(ctdt_ref)
    ci = {k: col_index(ch, label) for k, label in [
        ("manganh", "Mã ngành đào tạo"), ("tennganh", "Tên ngành đào tạo"),
        ("khoa", "Khoa/Viện")]}
    nganh_info: dict[str, tuple] = {}
    nganh_soct: dict[str, int] = {}
    for r in crows:
        raw = norm(r[ci["manganh"]]).split(".")[0] if ci["manganh"] is not None else ""
        mn = DIGITS_RE.sub("", raw)[:7]
        if len(mn) != 7:
            continue
        nganh_info.setdefault(mn, (norm(r[ci["tennganh"]]),
                                   normalize_unit(r[ci["khoa"]], canon)))
        nganh_soct[mn] = nganh_soct.get(mn, 0) + 1

    # HT_NCS: đếm theo mã ngành; leading 8 = ThS, 9 = TS
    nh, nrows = read_sheet(sdh_ref)
    ni_ma = col_index(nh, "Mã chương trình đào tạo")
    ni_tt = col_index(nh, "Trạng thái")
    ni_name = col_index(nh, "Họ và tên")
    counts: dict[str, dict] = {}
    for r in nrows:
        if ni_name is not None and norm(r[ni_name]).lower() == "mẫu":
            continue
        code = DIGITS_RE.sub("", norm(r[ni_ma]))[:7] if ni_ma is not None else ""
        if len(code) != 7:
            continue
        if SDH_ACTIVE_STATUS not in (norm(r[ni_tt]).lower() if ni_tt is not None else ""):
            continue
        base = code[1:]
        e = counts.setdefault(base, {"ths": 0, "ts": 0, "ma_ths": None, "ma_ts": None})
        if code[0] == "8":
            e["ths"] += 1
            e["ma_ths"] = code
        elif code[0] == "9":
            e["ts"] += 1
            e["ma_ts"] = code

    out = [['Stt', 'Đơn vị quản lý', 'Ngành', 'Mã_Ngành_ThS', 'Quy_mô_ThS',
            'Mã_Ngành_TS', 'Quy_mô_TS', 'Số chương trình đào tạo']]
    n = 0
    for _, e in sorted(counts.items(), key=lambda kv: -(kv[1]["ths"] + kv[1]["ts"])):
        info = nganh_info.get(e["ma_ths"]) or nganh_info.get(e["ma_ts"])
        if info is None:                       # ngành không có trong CTDT -> bỏ
            continue
        ten, khoa = info
        soct = nganh_soct.get(e["ma_ths"], 0) + nganh_soct.get(e["ma_ts"], 0)
        n += 1
        out.append([n, khoa, ten,
                    to_int(e["ma_ths"]), e["ths"] or None,
                    to_int(e["ma_ts"]), e["ts"] or None,
                    soct or None])
    return out, n


def build_map(ref: SourceRef):
    header, rows = read_sheet(ref)
    i_ma, i_dv = col_index(header, "Mã quản lý"), col_index(header, "Đơn vị giảng dạy")
    out = [['Mã quản lý', 'Đơn vị giảng dạy']]
    n = 0
    for r in rows:
        ma = to_int(r[i_ma]) if i_ma is not None else None
        dv = norm(r[i_dv]) if i_dv is not None else ""
        if ma is None or not dv:
            continue
        n += 1
        out.append([ma, dv])
    return out, n


def build_ctdt(ref: SourceRef, canon: dict):
    header, rows = read_sheet(ref)
    ci = {k: col_index(header, label) for k, label in [
        ("stt", "Stt"), ("mact", "Mã chương trình"), ("tenct", "Tên chương trình"),
        ("manganh", "Mã ngành đào tạo"), ("tennganh", "Tên ngành đào tạo"),
        ("linhvuc", "Lĩnh vực đào tạo"), ("truong", "Trường thành viên"),
        ("khoa", "Khoa/Viện"), ("nam", "Năm tuyển sinh"), ("trinhdo", "Trình độ"),
        ("loaihinh", "Loại hình"), ("hinhthuc", "Hình thức đào tạo")]}

    def g(r, key):
        i = ci[key]
        return r[i] if (i is not None and i < len(r)) else None

    # giữ đúng layout target: 1 cột trống D giữa "Tên chương trình" và "Mã ngành"
    out = [['Stt', 'Mã chương trình', 'Tên chương trình', None, 'Mã ngành đào tạo',
            'Tên ngành đào tạo', 'Lĩnh vực đào tạo', 'Trường thành viên', 'Khoa/Viện',
            'Năm tuyển sinh/dự kiến', 'Trình độ', 'Loại hình', 'Hình thức đào tạo', 'Ghi chú']]
    n = 0
    for r in rows:
        if not norm(g(r, "mact")) and not norm(g(r, "tenct")):
            continue
        n += 1
        out.append([
            to_int(g(r, "stt")) or n,
            norm(g(r, "mact")) or None, norm(g(r, "tenct")) or None, None,
            to_int(g(r, "manganh")), norm(g(r, "tennganh")) or None,
            norm(g(r, "linhvuc")) or None, norm(g(r, "truong")) or None,
            normalize_unit(g(r, "khoa"), canon), to_int(g(r, "nam")),
            norm(g(r, "trinhdo")) or None, norm(g(r, "loaihinh")) or None,
            norm(g(r, "hinhthuc")) or None, None,
        ])
    return out, n


# =========================================================
# GHI FILE GỘP
# =========================================================
def write_merged(sheets, path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for name, data in sheets:
        ws = wb.create_sheet(title=name[:31])
        for row in data:
            ws.append(list(row))
    try:
        wb.save(path)
    except PermissionError as e:
        raise PermissionError(
            f"Khong ghi duoc '{path}' (file dang mo trong Excel?)."
        ) from e


# =========================================================
# MAIN
# =========================================================
ROLE_LABEL = {
    "hr":          "Danh sách giảng viên      -> 'Quy mô giảng viên'",
    "demand":      "Danh sách môn học          -> 'Phụ lục 2. DS môn học 2025'",
    "course_unit": "Môn học -> Khoa phụ trách  (ghép vào Phụ lục 2)",
    "ug":          "Quy mô sinh viên ĐH        -> 'Quy mô sinh viên'",
    "ctdt":        "Chương trình đào tạo       -> 'Chuongtrinhdaotao'",
    "sdh":         "Học viên/NCS SĐH           -> 'Quy mô SĐH'",
    "map":         "Đơn vị giảng dạy           -> 'Đơn vị giảng dạy'",
}
REQUIRED_ROLES = list(ROLE_LABEL)


def run() -> None:
    print("=" * 64)
    print("  PIPELINE TUYỂN DỤNG GIẢNG VIÊN UEH")
    print("=" * 64)

    sources, unused, errors = discover_sources(DATA_DIR)

    print("\n[1] Nhận diện file nguồn:")
    for role in REQUIRED_ROLES:
        ref = sources.get(role)
        if ref:
            print(f"    [OK]  {ROLE_LABEL[role]}   ({ref.filename} / '{ref.sheet}')")
        else:
            print(f"    [!!]  THIEU: {ROLE_LABEL[role]}")
    for f in unused:
        print(f"    [--]  Bo qua (khong dung): {os.path.basename(f)}")
    for name, msg in errors:
        print(f"    [XX]  Loi doc file {name}: {msg}")

    missing = [r for r in REQUIRED_ROLES if r not in sources]
    if missing:
        raise ValueError(
            "Thieu du lieu cho: " + ", ".join(missing) +
            ". Vui long bo sung file tuong ung vao thu muc data/."
        )

    print("\n[2] Xu ly & lam sach...")
    canon = build_canonical_map(sources["course_unit"], sources["map"])
    print(f"    - Bang chuan hoa ten don vi: {len(canon)} muc")

    builders = [
        (SHEET_HR,     build_hr,     (sources["hr"],)),
        (SHEET_DEMAND, build_demand, (sources["demand"], sources["course_unit"], canon)),
        (SHEET_UG,     build_ug,     (sources["ug"], canon)),
        (SHEET_PG,     build_pg,     (sources["sdh"], sources["ctdt"], canon)),
        (SHEET_MAP,    build_map,    (sources["map"],)),
        (SHEET_CTDT,   build_ctdt,   (sources["ctdt"], canon)),
    ]
    sheets = []
    for sheet_name, fn, args in builders:
        try:
            data, n = fn(*args)
        except Exception as e:
            raise RuntimeError(f"Loi khi dung sheet '{sheet_name}': {e}") from e
        print(f"    - {sheet_name}: {n} dong")
        sheets.append((sheet_name, data))

    print(f"\n[3] Ghi file gop: {os.path.relpath(MERGED_PATH, BASE)}")
    os.makedirs(DATA_DIR, exist_ok=True)
    write_merged(sheets, MERGED_PATH)

    print("\n[4] Tinh dinh bien & xuat data.js / data.json cho index.html...")
    build_data.main(MERGED_PATH)

    print("\n" + "=" * 64)
    print("  HOAN TAT! Mo index.html de xem dashboard.")
    print("=" * 64)


def main() -> int:
    try:
        run()
        return 0
    except (FileNotFoundError, ValueError, PermissionError, RuntimeError) as e:
        print(f"\n[LOI] {e}", file=sys.stderr)
        return 1
    except KeyboardInterrupt:
        print("\n[HUY] Da dung pipeline.", file=sys.stderr)
        return 130


if __name__ == "__main__":
    sys.exit(main())
